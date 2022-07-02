import requests
from bs4 import BeautifulSoup
import os
import json
import sys
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages")
from requests_oauthlib import OAuth1Session
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/lib/python3.10/site-packages")
from dotenv import find_dotenv, load_dotenv
import requests
import schedule
import time
import openpyxl


def job():
    # .envファイルを探して読み込み
    env_file = find_dotenv()
    load_dotenv(env_file)  

    CONSUMER_KEY = os.environ.get('CONSUMER_KEY')
    CONSUMER_SECRET = os.environ.get('CONSUMER_SECRET')
    ACCESS_KEY = os.environ.get('ACCESS_KEY')
    ACCESS_KEY_SECRET = os.environ.get('ACCESS_KEY_SECRET')

    # Twitterの認証
    twitter = OAuth1Session(CONSUMER_KEY, CONSUMER_SECRET, ACCESS_KEY, ACCESS_KEY_SECRET)
    print(twitter)

    #エンドポイント
    url_text = 'https://api.twitter.com/1.1/statuses/update.json'
    url_media = "https://upload.twitter.com/1.1/media/upload.json"



    # ここまでTwitter投稿の準備
    print("投稿準備完了")



    URL = 'https://orenoraresne.com/'
    # リクエストヘッダの指定
    headers = {"User-Agent": "hoge"}
    response = requests.get(URL,  headers=headers)
    r_text=response.text
    soup = BeautifulSoup(r_text, 'html.parser')


    print('レアスニのページの取得完了')

    soup_article=soup.find_all("article",attrs={"class","post-list"})[0]
    soup_img=soup_article.find_all("img")[0]['src']
    soup_name=soup_article.find_all("h1",attrs={"class","entry-title"})[0].text
    soup_url=soup_article.find_all("a")[0]["href"]
    # 画像の処理
    response = requests.get(soup_img)
    image = response.content
    files = {"media" : image}
    req_media = twitter.post(url_media, files = files)
    media_id = json.loads(req_media.text)['media_id']
    print('画像の取得完了')
    try:
        print("ここから詳細ページ")
        URL = soup_url
        # リクエストヘッダの指定
        headers = {"User-Agent": "hoge"}
        response = requests.get(URL,  headers=headers)
        r_text=response.text
        soup = BeautifulSoup(r_text, 'html.parser')
        soup_p=soup.find_all("p",attrs={"class","has-text-align-center"})
        soup_result=""
        for i in range(len(soup_p)):
            soup_a=soup.find_all("p",attrs={"class","has-text-align-center"})[i].text
            if soup_a[:3]=="定価：":
                soup_purene=soup.find_all("p",attrs={"class","has-text-align-center"})[i].text.replace("俺的","")
            if soup_a[:2]=="結果":
                soup_result=soup.find_all("p",attrs={"class","has-text-align-center"})[i].text.replace("予想外れ","").replace("予想的中","")
        if soup_result=="":
            params = {'status':"予想プレ値!!!\n\n{}\n\n{}\n\n※某人気スニーカーブログ参照".format(soup_name,soup_purene),'media_ids':[media_id]}
            wb = openpyxl.load_workbook("/Users/kuramochiosuke/Desktop/プレ値/price.xlsx")
            ws = wb["Sheet1"]
            for i in range(wb['Sheet1'].max_row):
                if ws.cell(row=i+1,column=1).value==params["status"]:
                    print("投稿済みです")
                    break    
                elif i==wb['Sheet1'].max_row-1:   
                    ws.cell(row=wb['Sheet1'].max_row+1,column=1).value = params["status"]
                    wb.save('/Users/kuramochiosuke/Desktop/プレ値/price.xlsx')
                    print("保存しました")
                    twitter.post(url_text, params = params)
                    print("投稿しました") 
            print("処理終了")
            print("")
        else:
            soup_teika=soup.find_all("div",attrs={"class","cboxcomment"})[0].text[soup.find_all("div",attrs={"class","cboxcomment"})[0].text.find("定価"):].replace("\n","")    
            params = {'status':"結果発表!!!\n\n{}\n\n{}\n\n{}\n\n※某人気スニーカーブログ参照".format(soup_name,soup_teika,soup_result),'media_ids':[media_id]}
            params['status']
            wb = openpyxl.load_workbook("/Users/kuramochiosuke/Desktop/プレ値/price.xlsx")
            ws = wb["Sheet1"]
            for i in range(wb['Sheet1'].max_row):
                if ws.cell(row=i+1,column=1).value==params["status"]:
                    print("投稿済みです")
                    break    
                elif i==wb['Sheet1'].max_row-1:    
                    ws.cell(row=wb['Sheet1'].max_row+1,column=1).value = params["status"]
                    wb.save('/Users/kuramochiosuke/Desktop/プレ値/price.xlsx')
                    print("保存しました")
                    twitter.post(url_text, params = params)
                    print("投稿しました") 
            print("処理終了")
            print("")
    except IndexError:
        print("INDEX エラーです")       
        print("")
    
    except FileNotFoundError:
        print("NOT FILE エラーです")       
        print("")

    except KeyError:
        print("KeyError エラーです")       
        print("")



def main():
    schedule.every(1).minutes.do(job)
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == '__main__':
    main()
